import os
import pandas as pd
import xml.etree.ElementTree as et
import openpyxl as op
import sys
from openpyxl.styles import PatternFill
import time

dictionary_value_type_descriptions = {'0': ['Missing data', 'Отсутствие данных'],
                                      '1': ['Array of homogeneous data', 'Массив однородных данных'],
                                      '2': ['Structure from data of different types',
                                            'Структура из данных разных типов'],
                                      '3': ['Logical data (TRUE, FALSE)', 'Логические данные (TRUE, FALSE)'],
                                      '4': ['Bit sequence', 'Последовательность битов'],
                                      '5': ['32-bit signed integer', '32-разрядное целое со знаком'],
                                      '6': ['32-bit unsigned integer', '32-разрядное целое без знака'],
                                      '9': ['Byte sequence', 'Последовательность байтов'],
                                      '10': ['ASCII character sequence', 'Последовательность ASCII символов'],
                                      '12': ['UTF-8 character sequence', 'Последовательность символов UTF-8'],
                                      '13': ['BCD byte encoding', 'Двоично-десятичная кодировка байта'],
                                      '15': ['8-bit signed integer', '8-разрядное целое число со знаком'],
                                      '16': ['16-bit signed integer', '16-разрядное целое число со знаком'],
                                      '17': ['8-bit unsigned integer', '8-разрядное целое число без знака'],
                                      '18': ['16-bit unsigned integer', '16-разрядное целое число без знака'],
                                      '19': ['Packed data array', 'Массив упакованных данных'],
                                      '20': ['64-bit signed integer', '64-разрядное целое со знаком'],
                                      '21': ['64-bit unsigned integer', '64-разрядное целое без знака'],
                                      '22': ['Enumeration', 'Перечисление'],
                                      '23': ['4-byte string - floating point number',
                                             '4-байтовая строка - число с плавающей запятой'],
                                      '24': ['8-byte string - floating point number',
                                             '8-байтовая строка - число с плавающей запятой'],
                                      '25': ['12-byte datetime string', '12-байтовая строка дата-время'],
                                      '26': ['5-byte "Date" string', '5-байтовая строка «Дата»'],
                                      '27': ['4-byte string "Time', '4-байтовая строка «Время»']}


# function to get names for obis codes
def get_names_for_obis_codes(obis_codes, file_with_names, mode='not russian'):
    names_table = pd.read_excel(file_with_names)
    result_names = []
    if mode == 'russian':
        names_list = names_table['NameRus'].to_list()
    else:
        names_list = names_table['NameEng'].to_list()
    for obis in obis_codes:
        if obis in names_table['OBIS'].to_list():
            result_names.append(names_list[names_table['OBIS'].to_list().index(obis)])
        else:
            result_names.append('-')
    return result_names


# function to get unique obis codes from xml files
def get_unique_obis_codes(xml_files):
    result_obis_codes_list = []
    for xml_file in xml_files:
        root_node = et.parse(xml_file).getroot()
        all_tags = list(root_node.iter())
        for i in range(len(all_tags)):
            if all_tags[i].tag == 'LN':
                current_obis_code = all_tags[i].text
                if current_obis_code not in result_obis_codes_list:
                    result_obis_codes_list.append(current_obis_code)
    return result_obis_codes_list


# function to get xml info for obis codes
def get_xml_info_for_obis_codes(xml_files, unique_obis_codes):
    info_from_files = []
    for xml_file in xml_files:
        xml_info = []
        root_node = et.parse(xml_file).getroot()
        all_tags = list(root_node.iter())
        for obis_code in unique_obis_codes:
            is_found_obis = False
            current_obis_info = {'Description': None, 'Value Type(code)': None, 'Value Type(description)': None,
                                 'Scaler': None}
            for i in range(len(all_tags)):
                if all_tags[i].tag == 'LN' and all_tags[i].text == obis_code:
                    is_found_obis = True
                    parent = all_tags[i - 1]
                    for children in parent:
                        if children.tag == 'Description' and current_obis_info['Description'] is None:
                            current_obis_info['Description'] = children.text
                        elif children.tag == 'Value' and current_obis_info['Value Type(code)'] is None: # and children.find('Type')
                            current_obis_info['Value Type(code)'] = children.attrib['Type']
                            current_obis_info['Value Type(description)'] = \
                                dictionary_value_type_descriptions[children.attrib['Type']][0]
                        elif children.tag == 'Scaler' and current_obis_info['Scaler'] is None:
                            current_obis_info['Scaler'] = children.text
            for key, value in current_obis_info.items():
                if value is None:
                    current_obis_info[key] = '-'
            if not is_found_obis:
                current_obis_info['Description'] = 'Not found'
            xml_info.append(list(current_obis_info.values()))
        info_from_files.append(xml_info)
    return info_from_files


# function to find xml files in working directory
def find_xml_files(dir_name):
    xml_files = []
    for fileName in os.listdir(dir_name):
        if fileName.endswith("xml") or fileName.endswith("XML"):
            xml_files.append(fileName)
            print(fileName)
    if len(xml_files) > 1:
        return xml_files
    else:
        print('There is less then two xml files.')
        sys.exit()


# function to make columns for trees comparing table
def make_multi_columns_for_tree_comparing(xml_file):
    return pd.MultiIndex.from_product(
        [[xml_file], ['Description', 'Value Type (code)', 'Value Type (description)', 'Scaler']])


# function to compare two xml files
def make_comparison_between_xml_files(info_from_xml_files):
    first_xml_info = info_from_xml_files[0]
    second_xml_info = info_from_xml_files[1]
    result_comparison = []
    final_result = []
    for i in range(len(second_xml_info)):
        comparison = []
        if first_xml_info[i][0] != 'Not found' and second_xml_info[i][0] != 'Not found':
            for j in [3, 1]:
                if first_xml_info[i][j] == second_xml_info[i][j]:
                    comparison.append('Yes')
                else:
                    comparison.append('No')
        else:
            comparison.extend(['-', '-'])
        if comparison[0] == 'Yes' and comparison[1] == 'Yes':
            final_result.append('Success')
        elif first_xml_info[i][0] == 'Not found' or second_xml_info[i][0] == 'Not found':
            final_result.append('Fail')
        else:
            final_result.append('Fail')
        result_comparison.append(comparison)
    return result_comparison, final_result


# function to add color to comparison
def make_comparison_results_colour(file_name):
    workbook = op.load_workbook(file_name)
    for sheet in workbook:
        for i in range(4, len(list(sheet.rows)) + 1):
            for j in range(11, 14):
                if sheet.cell(i, column=j).value == 'Yes':
                    sheet.cell(i, column=j).font = op.styles.Font(color='008000', bold=True)
                elif sheet.cell(i, column=j).value == 'Success':
                    sheet.cell(i, column=j).fill = PatternFill('solid', fgColor='008000')
                elif sheet.cell(i, column=j).value == 'No' or sheet.cell(i, column=j).value == '-':
                    sheet.cell(i, column=j).font = op.styles.Font(color='FF0000', bold=True)
                else:
                    sheet.cell(i, column=j).fill = PatternFill('solid', fgColor='FF0000')
    workbook.save(file_name)


# function to get final Excel file
def get_result_xlsx_file(table_name, dir_name, xml_files):
    xml_file_first = xml_files[0].replace('.xml', '')
    xml_file_second = xml_files[1].replace('.xml', '')
    file_name = f'{xml_file_first}_' + f'{xml_file_second}_comparison.xlsx'
    try:
        writer = pd.ExcelWriter(dir_name + '/result_tables' + f'/{file_name}', engine="xlsxwriter")
        table_name.to_excel(writer, sheet_name='sheetName')
        worksheet = writer.sheets["sheetName"]
        worksheet.set_column('A:A', 10)
        worksheet.set_column('B:E', 25)
        worksheet.set_column('F:I', 25)
        worksheet.set_column('J:K', 15)
        worksheet.set_column('L:L', 10)
        writer.close()
    except PermissionError:
        print("File is busy, close it.")
    make_comparison_results_colour(dir_name + '/result_tables' + f'/{file_name}')
    make_sheets_with_profiles(xml_files, file_name, dir_name)
    create_sheet_modification_time_of_files(dir_name, xml_files, 'OBISNameRusEng.xlsx', file_name)
    print('Your file is ready.')


# function to get captured objects for profile
def get_captured_objects_from_profile(profile_obis_code, xml_files):
    dictionaries = []
    for xml_file in xml_files:
        root_node = et.parse(xml_file).getroot()
        objects_dictionary = {}
        for profile_tag in root_node.iter('GXDLMSProfileGeneric'):
            if profile_tag.find('LN').text == profile_obis_code:
                for tag in profile_tag:
                    if tag.tag == 'CaptureObjects':
                        index = 1
                        for item in tag:
                            obis_code = item.find('LN').text
                            objects_dictionary[obis_code] = [item.find('Attribute').text, index]
                            index += 1
                        break
        dictionaries.append(objects_dictionary)
    return dictionaries


# function to get all info for profiles
def make_profiles(dir_name, xml_files, mode='russian'):
    all_profile_info = {}
    names_table = pd.read_excel(dir_name + '/OBISNameRusEng.xlsx')
    obis_codes = names_table['OBIS'].to_list()
    if mode == 'russian':
        obis_names = names_table['NameRus'].to_list()
    else:
        obis_names = names_table['NameEng'].to_list()
    for file_name in os.listdir(dir_name + '/profiles'):
        profile_info_code_attr = []
        profile_info_comparing = []
        workbook = op.load_workbook(dir_name + '/profiles/' + file_name)
        profile_dictionaries = get_captured_objects_from_profile(file_name.replace('.xlsx', ''), xml_files)
        sheet = workbook.active
        max_row = sheet.max_row
        for i in range(2, max_row + 1):
            obis_info_code_attr = [sheet.cell(row=i, column=1).fill.fgColor.value]
            curr_obis_name = None
            for obis in obis_codes:
                if obis.strip() == sheet.cell(row=i, column=2).value.strip():
                    curr_obis_name = obis_names[obis_codes.index(obis)]
                    obis_info_code_attr.extend([sheet.cell(row=i, column=2).value, curr_obis_name,
                                                sheet.cell(row=i, column=4).value])
                    break
            if curr_obis_name is None:
                obis_info_code_attr.extend([sheet.cell(row=i, column=2).value, '-',
                                            sheet.cell(row=i, column=4).value])
            obis_info_comparing = []
            for dictionary in profile_dictionaries:
                if dictionary is not None and sheet.cell(row=i, column=2).value.strip() in dictionary.keys():
                    obis_info_comparing.append('Yes')
                    if int(dictionary[sheet.cell(row=i, column=2).value.strip()][0]) == sheet.cell(row=i,
                                                                                                   column=4).value:
                        obis_info_comparing.append('Yes')
                    else:
                        obis_info_comparing.append('No')
                    obis_info_comparing.append(dictionary[sheet.cell(row=i, column=2).value.strip()][1])
                    del dictionary[sheet.cell(row=i, column=2).value.strip()]
                else:
                    obis_info_comparing.extend(['Not found', '-', '-'])
            profile_info_code_attr.append(obis_info_code_attr)
            profile_info_comparing.append(obis_info_comparing)

        for dictionary in profile_dictionaries:
            if len(dictionary) != 0:
                for key, value in dictionary.items():
                    obis_info_code_attr = []
                    obis_info_comparing = []
                    curr_obis_name = None
                    is_found_obis = False
                    obis_found_index = None
                    for info in profile_info_code_attr:
                        if key == info[1]:
                            obis_found_index = profile_info_code_attr.index(info)
                            is_found_obis = True
                            break
                    if not is_found_obis:
                        obis_info_code_attr.extend(['red', key])
                        for obis in obis_codes:
                            if key == obis:
                                curr_obis_name = obis_names[obis_codes.index(obis)]
                                obis_info_code_attr.extend([obis_names[obis_codes.index(obis)], value[0]])
                                break
                        if curr_obis_name is None:
                            obis_info_code_attr.extend(['-', value[0]])
                        obis_info_comparing.extend([xml_files[profile_dictionaries.index(dictionary)], '-', '-', '-', '-', '-'])
                        profile_info_code_attr.append(obis_info_code_attr)
                        profile_info_comparing.append(obis_info_comparing)
                    else:
                        profile_info_comparing[obis_found_index][3] = xml_files[profile_dictionaries.index(dictionary)]
        all_profile_info[file_name.replace('.xlsx', '')] = [profile_info_code_attr, profile_info_comparing]
    return all_profile_info


# function to add profiles to Excel file
def make_sheets_with_profiles(xml_files, xlsx_file, dir_name, mode='russian'):
    all_profile_info = make_profiles(dir_name, xml_files)
    names_table = pd.read_excel(dir_name + '/OBISNameRusEng.xlsx')
    obis_codes = names_table['OBIS'].to_list()
    if mode == 'russian':
        obis_names = names_table['NameRus'].to_list()
    else:
        obis_names = names_table['NameEng'].to_list()
    for key, value in all_profile_info.items():
        curr_profile_name = None
        for obis in obis_codes:
            if key == obis:
                curr_profile_name = obis_names[obis_codes.index(obis)]
                break
        if curr_profile_name is None:
            curr_profile_name = '-'
        first_df = pd.DataFrame(value[0],
                                columns=pd.MultiIndex.from_product(
                                    [['Color', 'OBIS', f'Профиль: {curr_profile_name}', 'Attribute'], ['']]))
        del first_df['Color']
        columns_second_df = pd.MultiIndex.from_tuples(
            [(xml_files[0], 'OBIS(is found/not found)'), (xml_files[0], 'Attribute coincidence(yes/no)'), (xml_files[0], 'Index'),
             (xml_files[1], 'OBIS(is found/not found)'), (xml_files[1], 'Attribute coincidence(yes/no)'), (xml_files[1], 'Index')])
        second_df = pd.DataFrame(value[1], columns=columns_second_df)
        index_compare = []
        for compare_info in value[1]:
            if compare_info[2] != '-' and compare_info[5] != '-':
                if compare_info[2] == compare_info[5]:
                    index_compare.append('Yes')
                else:
                    index_compare.append('No')
            else:
                index_compare.append('-')
        columns_third_df = pd.MultiIndex.from_tuples([('', 'Index coincidence(yes/no)')])
        third_df = pd.DataFrame(index_compare, columns=columns_third_df)
        result_df = pd.concat([first_df, second_df, third_df], axis=1)
        result_df = result_df.set_index('OBIS')
        writer = pd.ExcelWriter(dir_name + '/result_tables' + f'/{xlsx_file}', engine="openpyxl", mode='a')
        result_df.to_excel(writer, sheet_name=f'{key}')
        worksheet = writer.sheets[key]
        max_row = worksheet.max_row
        worksheet.cell(1, 2).font = op.styles.Font(color='FF0000', bold=True)
        for i in range(4, max_row + 1):
            if value[0][i - 4][0] != '00000000' and value[0][i - 4][0] != 'red':
                for j in range(1, 10):
                    worksheet.cell(i, column=j).fill = PatternFill('solid', fgColor='ADD8E6')
            elif value[0][i - 4][0] == 'red':
                for j in range(1, 10):
                    worksheet.cell(i, column=j).fill = PatternFill('solid', fgColor='FF0000')
            if index_compare[i - 4] == 'Yes':
                worksheet.cell(i, column=10).fill = PatternFill('solid', fgColor='008000')
            elif index_compare[i - 4] == 'No':
                worksheet.cell(i, column=10).fill = PatternFill('solid', fgColor='FF0000')
        writer.close()


def create_sheet_modification_time_of_files(dir_name, xml_files, names_file, table_file):
    result_list = []
    for xml_file in xml_files:
        result_list.append([xml_file, time.ctime(os.path.getmtime(dir_name + '/' + xml_file))])
    result_list.append([names_file, time.ctime(os.path.getmtime(dir_name + '/' + names_file))])
    df = pd.DataFrame(result_list, columns=['File name', 'Modification time'])
    df = df.set_index('File name')
    writer = pd.ExcelWriter(dir_name + '/result_tables' + f'/{table_file}', engine="openpyxl", mode='a')
    df.to_excel(writer, sheet_name='Files versions')
    writer.close()


# function to get trees comparing table
def get_trees_comparing_table(dir_name):
    xml_files = find_xml_files(dir_name)
    unique_obis_codes = get_unique_obis_codes(xml_files)
    info_from_xml_files = get_xml_info_for_obis_codes(xml_files, unique_obis_codes)
    columns_first_sub_table = pd.MultiIndex.from_tuples([('OBIS-code', '')])
    first_sub_table = pd.DataFrame(unique_obis_codes, columns=columns_first_sub_table)
    obis_names = get_names_for_obis_codes(unique_obis_codes, dir_name + '/OBISNameRusEng.xlsx')
    columns_second_sub_table = pd.MultiIndex.from_tuples([('OBIS name', '')])
    second_sub_table = pd.DataFrame(obis_names, columns=columns_second_sub_table)
    columns_third_sub_table = make_multi_columns_for_tree_comparing(xml_files[0])
    columns_fourth_sub_table = make_multi_columns_for_tree_comparing(xml_files[1])
    third_sub_table = pd.DataFrame(info_from_xml_files[0], columns=columns_third_sub_table)
    fourth_sub_table = pd.DataFrame(info_from_xml_files[1], columns=columns_fourth_sub_table)
    comparison = make_comparison_between_xml_files(info_from_xml_files)
    yes_no_comparison = comparison[0]
    columns_fifth_sub_table = pd.MultiIndex.from_tuples([('Coincidence (yes/no)', 'Set yes/no by Scaler'),
                                                         ('Coincidence (yes/no)', 'Set yes/no by Value Type')])
    fifth_sub_table = pd.DataFrame(yes_no_comparison, columns=columns_fifth_sub_table)
    columns_sixth_sub_table = pd.MultiIndex.from_tuples([('Total', 'Success/Fail')])
    result_comparison = comparison[1]
    sixth_sub_table = pd.DataFrame(result_comparison, columns=columns_sixth_sub_table)
    table = pd.concat(
        [first_sub_table, second_sub_table, third_sub_table, fourth_sub_table, fifth_sub_table, sixth_sub_table],
        axis=1)
    table_without_index = table.set_index('OBIS-code')
    get_result_xlsx_file(table_without_index, dir_name, xml_files)
